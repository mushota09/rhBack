"""
Service d'export des données de paie vers Excel.
"""
import io
from typing import Dict, List, Optional, Any
from datetime import datetime, date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import QuerySet

from paie_app.models import periode_paie, entree_paie, retenue_employe
from paie_app.services.database_optimizer import DatabaseOptimizer
from user_app.models import employe


class ExportService:
    """Service pour exporter les données de paie vers Excel"""

    def __init__(self):
        self.db_optimizer = DatabaseOptimizer()

    def export_period_data(self, periode_id: int, export_type: str = 'excel') -> Dict[str, Any]:
        """
        Exporte les données d'une période de paie.

        Args:
            periode_id: ID de la période à exporter
            export_type: Type d'export ('excel', 'csv')

        Returns:
            Dict avec les informations du fichier exporté
        """
        try:
            periode = periode_paie.objects.get(id=periode_id)

            if export_type == 'excel':
                return self._export_period_to_excel(periode)
            elif export_type == 'csv':
                return self._export_period_to_csv(periode)
            else:
                raise ValueError(f"Type d'export non supporté: {export_type}")

        except periode_paie.DoesNotExist:
            raise ValueError(f"Période {periode_id} non trouvée")

    def _export_period_to_excel(self, periode: periode_paie) -> Dict[str, Any]:
        """Exporte une période vers Excel."""
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut

        # Créer les différentes feuilles
        self._create_summary_sheet(wb, periode)
        self._create_payroll_entries_sheet(wb, periode)
        self._create_deductions_sheet(wb, periode)
        self._create_statistics_sheet(wb, periode)

        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"paie_{periode.annee}_{periode.mois:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return {
            'filename': filename,
            'content': buffer.getvalue(),
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'size': len(buffer.getvalue())
        }
    def _create_summary_sheet(self, wb: Workbook, periode: periode_paie) -> None:
        """Crée la feuille de résumé de la période."""
        ws = wb.create_sheet("Résumé")

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=16)

        # Titre
        ws['A1'] = f"Résumé de Paie - {periode.annee}/{periode.mois:02d}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Informations générales
        row = 3
        info_data = [
            ("Période", f"{periode.annee}/{periode.mois:02d}"),
            ("Statut", periode.get_statut_display()),
            ("Nombre d'employés", str(periode.nombre_employes or 0)),
            ("Traité par", periode.traite_par.get_full_name() if periode.traite_par else 'N/A'),
            ("Date de traitement", periode.date_traitement.strftime('%d/%m/%Y') if periode.date_traitement else 'N/A'),
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = value
            row += 1

        # Informations financières
        row += 1
        financial_data = [
            ("Masse salariale brute", f"{periode.masse_salariale_brute:,.2f} USD"),
            ("Total cotisations patronales", f"{periode.total_cotisations_patronales:,.2f} USD"),
            ("Total cotisations salariales", f"{periode.total_cotisations_salariales:,.2f} USD"),
            ("Total net à payer", f"{periode.total_net_a_payer:,.2f} USD"),
        ]

        for label, value in financial_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = value
            row += 1

        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_payroll_entries_sheet(self, wb: Workbook, periode: periode_paie) -> None:
        """Crée la feuille des entrées de paie."""
        ws = wb.create_sheet("Entrées de Paie")

        # En-têtes
        headers = [
            "Nom", "Prénom", "Email", "Salaire Base", "Indemnité Logement",
            "Indemnité Déplacement", "Indemnité Fonction", "Allocation Familiale",
            "Autres Avantages", "Salaire Brut", "Total Charges", "Base Imposable",
            "Salaire Net", "Bulletin Généré", "Date Calcul"
        ]

        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill =PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Récupérer les données avec requêtes optimisées
        entries = self.db_optimizer.get_optimized_entree_paie_queryset(periode.id)

        # Remplir les données
        for row, entry in enumerate(entries, 2):
            ws.cell(row=row, column=1, value=entry.employe_id.nom)
            ws.cell(row=row, column=2, value=entry.employe_id.prenom)
            ws.cell(row=row, column=3, value=entry.employe_id.email_personnel)
            ws.cell(row=row, column=4, value=float(entry.salaire_base))
            ws.cell(row=row, column=5, value=float(entry.indemnite_logement))
            ws.cell(row=row, column=6, value=float(entry.indemnite_deplacement))
            ws.cell(row=row, column=7, value=float(entry.indemnite_fonction))
            ws.cell(row=row, column=8, value=float(entry.allocation_familiale))
            ws.cell(row=row, column=9, value=float(entry.autres_avantages))
            ws.cell(row=row, column=10, value=float(entry.salaire_brut))
            ws.cell(row=row, column=11, value=float(entry.total_charge_salariale))
            ws.cell(row=row, column=12, value=float(entry.base_imposable))
            ws.cell(row=row, column=13, value=float(entry.salaire_net))
            ws.cell(row=row, column=14, value="Oui" if entry.payslip_generated else "Non")
            ws.cell(row=row, column=15, value=entry.calculated_at.strftime('%d/%m/%Y %H:%M') if entry.calculated_at else 'N/A')

        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Ajouter des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
    def _create_deductions_sheet(self, wb: Workbook, periode: periode_paie) -> None:
        """Crée la feuille des retenues."""
        ws = wb.create_sheet("Retenues")

        # En-têtes
        headers = [
            "Employé", "Type Retenue", "Description", "Montant Mensuel",
            "Montant Total", "Montant Déjà Déduit", "Date Début", "Date Fin",
            "Statut", "Récurrente"
        ]

        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Récupérer les retenues actives
        deductions = retenue_employe.objects.select_related('employe_id').filter(
            est_active=True
        ).order_by('employe_id__nom', 'type_retenue')

        # Remplir les données
        for row, deduction in enumerate(deductions, 2):
            ws.cell(row=row, column=1, value=f"{deduction.employe_id.nom} {deduction.employe_id.prenom}")
            ws.cell(row=row, column=2, value=deduction.get_type_retenue_display())
            ws.cell(row=row, column=3, value=deduction.description)
            ws.cell(row=row, column=4, value=float(deduction.montant_mensuel))
            ws.cell(row=row, column=5, value=float(deduction.montant_total) if deduction.montant_total else 0)
            ws.cell(row=row, column=6, value=float(deduction.montant_deja_deduit))
            ws.cell(row=row, column=7, value=deduction.date_debut.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=8, value=deduction.date_fin.strftime('%d/%m/%Y') if deduction.date_fin else 'N/A')
            ws.cell(row=row, column=9, value="Active" if deduction.est_active else "Inactive")
            ws.cell(row=row, column=10, value="Oui" if deduction.est_recurrente else "Non")

        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_statistics_sheet(self, wb: Workbook, periode: periode_paie) -> None:
        """Crée la feuille des statistiques."""
        ws = wb.create_sheet("Statistiques")

        # Récupérer les statistiques avec le cache
        stats = self.db_optimizer.get_period_statistics_cached(periode.id)

        # Titre
        ws['A1'] = "Statistiques de la Période"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')

        # Statistiques générales
        row = 3
        general_stats = [
            ("Nombre d'employés", str(stats.get('total_employees', 0))),
            ("Total salaire brut", f"{stats.get('total_salaire_brut', 0) or 0:,.2f} USD"),
            ("Total des charges", f"{stats.get('total_charges', 0) or 0:,.2f} USD"),
            ("Salaire brut moyen", f"{stats.get('avg_salaire_brut', 0) or 0:,.2f} USD"),
            ("Salaire net moyen", f"{stats.get('avg_salaire_net', 0) or 0:,.2f} USD"),
        ]

        header_font = Font(bold=True)
        for label, value in general_stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = value
            row += 1

        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    def _export_period_to_csv(self, periode: periode_paie) -> Dict[str, Any]:
        """Exporte une période vers CSV (implémentation simplifiée)."""
        import csv

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # En-têtes
        writer.writerow([
            "Nom", "Prénom", "Email", "Salaire Base", "Salaire Brut", "Salaire Net"
        ])

        # Données
        entries = entree_paie.objects.select_related('employe_id').filter(
            periode_paie_id=periode.id
        )

        for entry in entries:
            writer.writerow([
                entry.employe_id.nom,
                entry.employe_id.prenom,
                entry.employe_id.email_personnel,
                float(entry.salaire_base),
                float(entry.salaire_brut),
                float(entry.salaire_net)
            ])

        filename = f"paie_{periode.annee}_{periode.mois:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return {
            'filename': filename,
            'content': buffer.getvalue().encode('utf-8'),
            'content_type': 'text/csv',
            'size': len(buffer.getvalue().encode('utf-8'))
        }

    def export_employees_data(self, format_type: str = 'excel') -> Dict[str, Any]:
        """
        Exporte les données des employés.

        Args:
            format_type: Type de format ('excel', 'csv')

        Returns:
            Dict avec les informations du fichier exporté
        """
        if format_type == 'excel':
            return self._export_employees_to_excel()
        elif format_type == 'csv':
            return self._export_employees_to_csv()
        else:
            raise ValueError(f"Format non supporté: {format_type}")
    def _export_employees_to_excel(self) -> Dict[str, Any]:
        """Exporte les employés vers Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Employés"

        # En-têtes
        headers = [
            "Nom", "Prénom", "Email", "Date Naissance", "Date Embauche",
            "Sexe", "Statut Matrimonial", "Statut Emploi", "Nationalité",
            "Banque", "Numéro Compte", "Niveau Étude", "Numéro INSS",
            "Téléphone", "Adresse", "Nombre Enfants"
        ]

        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Récupérer les employés avec requêtes optimisées
        employees = self.db_optimizer.get_optimized_employe_queryset()

        # Remplir les données
        for row, employee in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=employee.nom)
            ws.cell(row=row, column=2, value=employee.prenom)
            ws.cell(row=row, column=3, value=employee.email_personnel)
            ws.cell(row=row, column=4, value=employee.date_naissance.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=5, value=employee.date_embauche.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=6, value=employee.get_sexe_display())
            ws.cell(row=row, column=7, value=employee.get_statut_matrimonial_display())
            ws.cell(row=row, column=8, value=employee.get_statut_emploi_display())
            ws.cell(row=row, column=9, value=employee.nationalite)
            ws.cell(row=row, column=10, value=employee.banque)
            ws.cell(row=row, column=11, value=employee.numero_compte)
            ws.cell(row=row, column=12, value=employee.niveau_etude or 'N/A')
            ws.cell(row=row, column=13, value=employee.numero_inss)
            ws.cell(row=row, column=14, value=employee.telephone_personnel)
            ws.cell(row=row, column=15, value=employee.adresse_ligne1)
            ws.cell(row=row, column=16, value=employee.nombre_enfants)

        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"employes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return {
            'filename': filename,
            'content': buffer.getvalue(),
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'size': len(buffer.getvalue())
        }
    def _export_employees_to_csv(self) -> Dict[str, Any]:
        """Exporte les employés vers CSV."""
        import csv

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # En-têtes
        writer.writerow([
            "Nom", "Prénom", "Email", "Date Naissance", "Date Embauche",
            "Statut Emploi", "Téléphone", "Nombre Enfants"
        ])

        # Données
        employees = employe.objects.filter(statut_emploi='ACTIVE')

        for employee in employees:
            writer.writerow([
                employee.nom,
                employee.prenom,
                employee.email_personnel,
                employee.date_naissance.strftime('%d/%m/%Y'),
                employee.date_embauche.strftime('%d/%m/%Y'),
                employee.get_statut_emploi_display(),
                employee.telephone_personnel,
                employee.nombre_enfants
            ])

        filename = f"employes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return {
            'filename': filename,
            'content': buffer.getvalue().encode('utf-8'),
            'content_type': 'text/csv',
            'size': len(buffer.getvalue().encode('utf-8'))
        }

    def create_http_response(self, export_data: Dict[str, Any]) -> HttpResponse:
        """
        Crée une réponse HTTP pour télécharger le fichier exporté.

        Args:
            export_data: Données d'export retournées par les méthodes d'export

        Returns:
            HttpResponse configurée pour le téléchargement
        """
        response = HttpResponse(
            export_data['content'],
            content_type=export_data['content_type']
        )
        response['Content-Disposition'] = f'attachment; filename="{export_data["filename"]}"'
        response['Content-Length'] = export_data['size']

        return response
