"""
Vues API pour les périodes de paie.
"""
from decimal import Decimal
from io import BytesIO
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from adrf.viewsets import ModelViewSet
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
import openpyxl
from openpyxl.styles import Font, Alignment

from paie_app.models import periode_paie, entree_paie
from paie_app.services import PeriodProcessorService
from utilities.permissions import (
    CanProcessPayroll, CanApprovePayroll,
    CanExportData, PayrollReadPermission, PayrollWritePermission
)
from utilities.decorators import audit_action
from paie_app.modules.periode_paie.serializers import (
    PeriodePaieSerializer,
    PeriodePaieListSerializer,
    PeriodePaieProcessSerializer,
    PeriodePaieApprovalSerializer
)


class PeriodePaieAPIView(ModelViewSet):
    """API View pour la gestion des périodes de paie."""

    queryset = periode_paie.objects.all().order_by('-created_at', '-mois')
    serializer_class = PeriodePaieSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['annee', 'mois', 'statut']
    search_fields = ['annee', 'mois']
    ordering_fields = ['annee', 'mois', 'created_at', 'statut']
    ordering = ['-annee', '-mois']

    # Role-based permissions
    permission_classes = [IsAuthenticated]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.period_processor = PeriodProcessorService()

    def get_permissions(self):
        """
        Instantiate and return the list of permissions required for this view.
        """
        if self.action in ['list', 'retrieve', 'statistics']:
            # Read operations - allow HR managers and employees
            permission_classes = [PayrollReadPermission]
        elif self.action in ['create', 'update', 'partial_update', 'destroy']:
            # Write operations - only HR admins
            permission_classes = [PayrollWritePermission]
        elif self.action in ['process_period', 'finalize_period']:
            # Processing operations - only users who can process payroll
            permission_classes = [CanProcessPayroll]
        elif self.action == 'approve_period':
            # Approval operations - only users who can approve payroll
            permission_classes = [CanApprovePayroll]
        elif self.action == 'export_excel':
            # Export operations - only users who can export data
            permission_classes = [CanExportData]
        else:
            # Default to authenticated users
            permission_classes = [IsAuthenticated]

        return [permission() for permission in permission_classes]

    def get_serializer_class(self):
        """Retourne le serializer approprié selon l'action."""
        if self.action == 'list':
            return PeriodePaieListSerializer
        if self.action == 'process_period':
            return PeriodePaieProcessSerializer
        if self.action == 'approve_period':
            return PeriodePaieApprovalSerializer
        return PeriodePaieSerializer

    async def list(self, request, *args, **kwargs):
        """Liste les périodes de paie avec filtres."""
        return await super().list(request, *args, **kwargs)

    @audit_action('CREATE', 'periode_paie')
    async def create(self, request, *args, **kwargs):
        """Crée une nouvelle période de paie."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Utiliser le service pour créer la période avec calcul automatique
        try:
            periode = await self.period_processor.create_period(
                annee=serializer.validated_data['annee'],
                mois=serializer.validated_data['mois'],
                user_id=request.user.id
            )

            # Sérialiser la réponse
            response_serializer = self.get_serializer(periode)
            return Response(
                await response_serializer.adata,
                status=status.HTTP_201_CREATED
            )

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    async def retrieve(self, request, *args, **kwargs):
        """Récupère les détails d'une période de paie."""
        return await super().retrieve(request, *args, **kwargs)

    @audit_action('UPDATE', 'periode_paie')
    async def update(self, request, *args, **kwargs):
        """Met à jour une période de paie."""
        partial = kwargs.pop('partial', False)
        instance = await self.aget_object()

        # Vérifier que la période n'est pas approuvée
        if instance.statut == 'APPROVED':
            return Response(
                {'error': 'Impossible de modifier une période approuvée'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = self.get_serializer(
            instance, data=request.data, partial=partial
        )
        serializer.is_valid(raise_exception=True)
        await serializer.asave()

        return Response(await serializer.adata)

    @audit_action('DELETE', 'periode_paie')
    async def destroy(self, request, *args, **kwargs):
        """Supprime une période de paie."""
        instance = await self.aget_object()

        # Vérifier que la période peut être supprimée
        if instance.statut in ['FINALIZED', 'APPROVED']:
            return Response(
                {'error': 'Impossible de supprimer une période finalisée ou approuvée'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérifier s'il y a des entrées de paie
        entrees_count = await entree_paie.objects.filter(
            periode_paie_id=instance
        ).acount()

        if entrees_count > 0:
            return Response(
                {'error': f'Impossible de supprimer: {entrees_count} entrées de paie associées'},
                status=status.HTTP_400_BAD_REQUEST
            )

        await instance.adelete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    @audit_action('PROCESS', 'periode_paie')
    async def process_period(self, request, pk=None):
        """Traite une période de paie (calcul des salaires)."""
        periode = await self.aget_object()

        serializer = self.get_serializer(
            data=request.data,
            context={'periode': periode}
        )
        serializer.is_valid(raise_exception=True)

        try:
            results = await self.period_processor.process_period(periode.id)

            return Response({
                'success': True,
                'message': 'Période traitée avec succès',
                'results': results
            })

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    async def finalize_period(self, request, pk=None):
        """Finalise une période de paie."""
        periode = await self.aget_object()

        try:
            success = await self.period_processor.finalize_period(periode.id)

            if success:
                return Response({
                    'success': True,
                    'message': 'Période finalisée avec succès'
                })
            else:
                return Response(
                    {'error': 'Échec de la finalisation'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    @audit_action('APPROVE', 'periode_paie')
    async def approve_period(self, request, pk=None):
        """Approuve une période de paie."""
        periode = await self.aget_object()

        serializer = self.get_serializer(
            data=request.data,
            context={'periode': periode}
        )
        serializer.is_valid(raise_exception=True)

        try:
            success = await self.period_processor.approve_period(
                periode.id, request.user.id
            )

            if success:
                return Response({
                    'success': True,
                    'message': 'Période approuvée avec succès'
                })
            else:
                return Response(
                    {'error': 'Échec de l\'approbation'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'])
    async def validate_period(self, request, pk=None):
        """Valide une période de paie et retourne les erreurs."""
        periode = await self.aget_object()

        try:
            errors = await self.period_processor.validate_period(periode.id)

            return Response({
                'valid': len(errors) == 0,
                'errors': errors,
                'periode_id': periode.id
            })

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'])
    @audit_action('EXPORT', 'periode_paie')
    async def export_excel(self, request, pk=None):
        """Exporte une période de paie vers Excel."""
        periode = await self.aget_object()

        try:
            # Créer le workbook Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Paie_{periode.annee}_{periode.mois:02d}"

            # En-têtes
            headers = [
                'Employé', 'Email', 'N° INSS', 'Salaire Base',
                'Indemnité Logement', 'Allocation Familiale',
                'Salaire Brut', 'Cotisations', 'Salaire Net'
            ]

            # Style des en-têtes
            header_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = center_alignment

            # Récupérer les entrées de paie
            entrees = []
            async for entree in entree_paie.objects.filter(
                periode_paie_id=periode
            ).select_related('employe_id'):
                entrees.append(entree)

            # Données
            for row, entree in enumerate(entrees, 2):
                employe = entree.employe_id
                ws.cell(row=row, column=1, value=f"{employe.nom} {employe.prenom}")
                ws.cell(row=row, column=2, value=employe.email_personnel)
                ws.cell(row=row, column=3, value=employe.numero_inss)
                ws.cell(row=row, column=4, value=float(entree.salaire_base))
                ws.cell(row=row, column=5, value=float(entree.indemnite_logement or 0))
                ws.cell(row=row, column=6, value=float(entree.allocation_familiale or 0))
                ws.cell(row=row, column=7, value=float(entree.salaire_brut))
                ws.cell(row=row, column=8, value=float(entree.cotisations_salariales or 0))
                ws.cell(row=row, column=9, value=float(entree.salaire_net))

            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Sauvegarder dans un buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Créer la réponse HTTP
            filename = f"paie_{periode.annee}_{periode.mois:02d}.xlsx"
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {'error': f'Erreur lors de l\'export: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    async def statistics(self, request):
        """Retourne les statistiques des périodes de paie."""
        try:
            # Compter les périodes par statut
            stats = {}
            async for periode in periode_paie.objects.all():
                statut = periode.statut
                if statut not in stats:
                    stats[statut] = 0
                stats[statut] += 1

            # Calculer les totaux
            total_periodes = sum(stats.values())
            total_masse_salariale = Decimal('0')
            total_net_a_payer = Decimal('0')

            async for periode in periode_paie.objects.all():
                if periode.masse_salariale_brute:
                    total_masse_salariale += periode.masse_salariale_brute
                if periode.total_net_a_payer:
                    total_net_a_payer += periode.total_net_a_payer

            return Response({
                'total_periodes': total_periodes,
                'periodes_par_statut': stats,
                'total_masse_salariale': total_masse_salariale,
                'total_net_a_payer': total_net_a_payer
            })

        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul des statistiques: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
