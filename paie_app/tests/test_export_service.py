"""
Tests unitaires pour le service d'export.
"""
from decimal import Decimal
from datetime import date
from django.test import TestCase
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
import io

from paie_app.services.export_service import ExportService
from paie_app.models import periode_paie, entree_paie, retenue_employe
from user_app.models import employe, contrat

User = get_user_model()


class ExportServiceTests(TestCase):
    """Tests unitaires pour ExportService"""

    def setUp(self):
        """Configuration des tests"""
        self.service = ExportService()

        # Créer un utilisateur de test
        self.user = User.objects.create(
            email='test@example.com',
            nom='Test',
            prenom='User',
            is_active=True
        )
        self.user.set_password('testpass123')
        self.user.save()

        # Créer une période de test
        self.periode = periode_paie.objects.create(
            annee=2024,
            mois=1,
            traite_par=self.user,
            statut='COMPLETED',
            nombre_employes=2,
            masse_salariale_brute=Decimal('1200000'),
            total_net_a_payer=Decimal('900000')
        )

        # Créer des employés de test
        self.employe1 = employe.objects.create(
            email_personnel='employe1@example.com',
            email_professionnel='employe1.pro@example.com',
            nom='Dupont',
            prenom='Jean',
            date_naissance=date(1990, 1, 1),
            date_embauche=date(2020, 1, 1),
            sexe='M',
            statut_matrimonial='S',
            statut_emploi='ACTIVE',
            nationalite='Congolaise',
            banque='Test Bank',
            numero_compte='123456789',
            niveau_etude='Universitaire',
            numero_inss='123456789',
            telephone_personnel='123456789',
            adresse_ligne1='Test Address',
            nombre_enfants=1,
            nom_contact_urgence='Contact',
            lien_contact_urgence='Parent',
            telephone_contact_urgence='987654321'
        )

        self.employe2 = employe.objects.create(
            email_personnel='employe2@example.com',
            email_professionnel='employe2.pro@example.com',
            nom='Martin',
            prenom='Marie',
            date_naissance=date(1985, 5, 15),
            date_embauche=date(2019, 3, 1),
            sexe='F',
            statut_matrimonial='M',
            statut_emploi='ACTIVE',
            nationalite='Congolaise',
            banque='Test Bank 2',
            numero_compte='987654321',
            niveau_etude='Universitaire',
            numero_inss='987654321',
            telephone_personnel='987654321',
            adresse_ligne1='Test Address 2',
            nombre_enfants=2,
            nom_contact_urgence='Contact 2',
            lien_contact_urgence='Spouse',
            telephone_contact_urgence='123456789'
        )

        # Créer des entrées de paie de test
        self.entree1 = entree_paie.objects.create(
            employe_id=self.employe1,
            periode_paie_id=self.periode,
            salaire_base=Decimal('500000'),
            indemnite_logement=Decimal('50000'),
            indemnite_deplacement=Decimal('25000'),
            indemnite_fonction=Decimal('30000'),
            allocation_familiale=Decimal('5000'),
            autres_avantages=Decimal('10000'),
            salaire_brut=Decimal('620000'),
            total_charge_salariale=Decimal('650000'),
            base_imposable=Decimal('600000'),
            salaire_net=Decimal('450000'),
            payslip_generated=True
        )

        self.entree2 = entree_paie.objects.create(
            employe_id=self.employe2,
            periode_paie_id=self.periode,
            salaire_base=Decimal('600000'),
            indemnite_logement=Decimal('60000'),
            indemnite_deplacement=Decimal('30000'),
            indemnite_fonction=Decimal('40000'),
            allocation_familiale=Decimal('10000'),
            autres_avantages=Decimal('15000'),
            salaire_brut=Decimal('755000'),
            total_charge_salariale=Decimal('800000'),
            base_imposable=Decimal('700000'),
            salaire_net=Decimal('550000'),
            payslip_generated=False
        )

        # Créer des retenues de test
        self.retenue1 = retenue_employe.objects.create(
            employe_id=self.employe1,
            type_retenue='LOAN',
            description='Prêt personnel',
            montant_mensuel=Decimal('50000'),
            montant_total=Decimal('500000'),
            montant_deja_deduit=Decimal('100000'),
            date_debut=date(2024, 1, 1),
            est_active=True,
            est_recurrente=True,
            cree_par=self.user
        )

    def test_export_period_to_excel(self):
        """Test l'export d'une période vers Excel"""
        result = self.service.export_period_data(self.periode.id, 'excel')

        # Vérifier les métadonnées du fichier
        self.assertIn('filename', result)
        self.assertIn('content', result)
        self.assertIn('content_type', result)
        self.assertIn('size', result)

        self.assertTrue(result['filename'].endswith('.xlsx'))
        self.assertEqual(result['content_type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertGreater(result['size'], 0)

        # Vérifier le contenu du fichier Excel
        wb = load_workbook(io.BytesIO(result['content']))

        # Vérifier que toutes les feuilles sont présentes
        expected_sheets = ['Résumé', 'Entrées de Paie', 'Retenues', 'Statistiques']
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, wb.sheetnames)

        # Vérifier le contenu de la feuille "Entrées de Paie"
        ws_entries = wb['Entrées de Paie']

        # Vérifier les en-têtes
        headers = [cell.value for cell in ws_entries[1]]
        self.assertIn('Nom', headers)
        self.assertIn('Prénom', headers)
        self.assertIn('Salaire Base', headers)
        self.assertIn('Salaire Net', headers)

        # Vérifier que les données des employés sont présentes
        employee_names = [ws_entries.cell(row=row, column=1).value for row in range(2, ws_entries.max_row + 1)]
        self.assertIn('Dupont', employee_names)
        self.assertIn('Martin', employee_names)

    def test_export_period_to_csv(self):
        """Test l'export d'une période vers CSV"""
        result = self.service.export_period_data(self.periode.id, 'csv')

        # Vérifier les métadonnées du fichier
        self.assertIn('filename', result)
        self.assertIn('content', result)
        self.assertIn('content_type', result)
        self.assertIn('size', result)

        self.assertTrue(result['filename'].endswith('.csv'))
        self.assertEqual(result['content_type'], 'text/csv')
        self.assertGreater(result['size'], 0)

        # Vérifier le contenu CSV
        content = result['content'].decode('utf-8')
        lines = content.strip().split('\n')

        # Vérifier les en-têtes
        headers = lines[0].split(',')
        self.assertIn('Nom', headers)
        self.assertIn('Prénom', headers)

        # Vérifier que les données sont présentes
        self.assertGreater(len(lines), 1)  # Au moins une ligne de données
        self.assertIn('Dupont', content)
        self.assertIn('Martin', content)

    def test_export_employees_to_excel(self):
        """Test l'export des employés vers Excel"""
        result = self.service.export_employees_data('excel')

        # Vérifier les métadonnées du fichier
        self.assertIn('filename', result)
        self.assertIn('content', result)
        self.assertIn('content_type', result)
        self.assertIn('size', result)

        self.assertTrue(result['filename'].startswith('employes_'))
        self.assertTrue(result['filename'].endswith('.xlsx'))
        self.assertEqual(result['content_type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Vérifier le contenu du fichier Excel
        wb = load_workbook(io.BytesIO(result['content']))
        ws = wb.active

        # Vérifier les en-têtes
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Nom', headers)
        self.assertIn('Prénom', headers)
        self.assertIn('Email', headers)
        self.assertIn('Statut Emploi', headers)

        # Vérifier que les données des employés sont présentes
        employee_names = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
        self.assertIn('Dupont', employee_names)
        self.assertIn('Martin', employee_names)

    def test_export_employees_to_csv(self):
        """Test l'export des employés vers CSV"""
        result = self.service.export_employees_data('csv')

        # Vérifier les métadonnées du fichier
        self.assertTrue(result['filename'].startswith('employes_'))
        self.assertTrue(result['filename'].endswith('.csv'))
        self.assertEqual(result['content_type'], 'text/csv')

        # Vérifier le contenu CSV
        content = result['content'].decode('utf-8')
        lines = content.strip().split('\n')

        # Vérifier les en-têtes
        headers = lines[0].split(',')
        self.assertIn('Nom', headers)
        self.assertIn('Prénom', headers)

        # Vérifier que les données sont présentes
        self.assertIn('Dupont', content)
        self.assertIn('Martin', content)

    def test_export_nonexistent_period(self):
        """Test l'export d'une période inexistante"""
        with self.assertRaises(ValueError) as context:
            self.service.export_period_data(99999, 'excel')

        self.assertIn('Période 99999 non trouvée', str(context.exception))

    def test_export_unsupported_format(self):
        """Test l'export avec un format non supporté"""
        with self.assertRaises(ValueError) as context:
            self.service.export_period_data(self.periode.id, 'pdf')

        self.assertIn('Type d\'export non supporté: pdf', str(context.exception))

    def test_create_http_response(self):
        """Test la création d'une réponse HTTP pour téléchargement"""
        export_data = {
            'filename': 'test.xlsx',
            'content': b'test content',
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'size': 12
        }

        response = self.service.create_http_response(export_data)

        # Vérifier les en-têtes de la réponse
        self.assertEqual(response['Content-Type'], export_data['content_type'])
        self.assertEqual(response['Content-Disposition'], 'attachment; filename="test.xlsx"')
        self.assertEqual(response['Content-Length'], '12')

        # Vérifier le contenu
        self.assertEqual(response.content, b'test content')

    def test_excel_file_structure(self):
        """Test la structure détaillée du fichier Excel"""
        result = self.service.export_period_data(self.periode.id, 'excel')
        wb = load_workbook(io.BytesIO(result['content']))

        # Test de la feuille Résumé
        ws_summary = wb['Résumé']
        self.assertEqual(ws_summary['A1'].value, f"Résumé de Paie - {self.periode.annee}/{self.periode.mois:02d}")

        # Test de la feuille Entrées de Paie
        ws_entries = wb['Entrées de Paie']

        # Vérifier que les données numériques sont correctes
        # Trouver la ligne de Jean Dupont
        jean_row = None
        for row in range(2, ws_entries.max_row + 1):
            if ws_entries.cell(row=row, column=1).value == 'Dupont':
                jean_row = row
                break

        self.assertIsNotNone(jean_row)
        self.assertEqual(ws_entries.cell(row=jean_row, column=2).value, 'Jean')  # Prénom
        self.assertEqual(ws_entries.cell(row=jean_row, column=4).value, 500000.0)  # Salaire Base
        self.assertEqual(ws_entries.cell(row=jean_row, column=13).value, 450000.0)  # Salaire Net

        # Test de la feuille Retenues
        ws_deductions = wb['Retenues']

        # Vérifier que la retenue est présente
        retenue_found = False
        for row in range(2, ws_deductions.max_row + 1):
            if ws_deductions.cell(row=row, column=3).value == 'Prêt personnel':
                retenue_found = True
                self.assertEqual(ws_deductions.cell(row=row, column=4).value, 50000.0)  # Montant mensuel
                break

        self.assertTrue(retenue_found, "La retenue de test n'a pas été trouvée dans l'export")

    def test_empty_period_export(self):
        """Test l'export d'une période sans entrées"""
        # Créer une période vide
        periode_vide = periode_paie.objects.create(
            annee=2024,
            mois=2,
            traite_par=self.user,
            statut='DRAFT'
        )

        result = self.service.export_period_data(periode_vide.id, 'excel')

        # Vérifier que l'export fonctionne même sans données
        self.assertIn('filename', result)
        self.assertGreater(result['size'], 0)

        # Vérifier le contenu
        wb = load_workbook(io.BytesIO(result['content']))
        ws_entries = wb['Entrées de Paie']

        # Seule la ligne d'en-tête devrait être présente
        self.assertEqual(ws_entries.max_row, 1)
